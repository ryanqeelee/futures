"""
导出工具模块

提供多种格式的数据导出功能，包括Excel、CSV、PDF报告生成
以及高级数据处理和格式化功能
"""

import logging
from datetime import datetime
from typing import Dict, List, Any, Optional, Union, IO
from pathlib import Path
import tempfile
import zipfile
import io
import base64

import streamlit as st
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import ScatterChart, Reference, LineChart, BarChart, PieChart
from openpyxl.chart.axis import DateAxis
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows

# PDF生成
try:
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
    from reportlab.platypus import PageBreak, KeepTogether
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch, cm
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.linecharts import HorizontalLineChart
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from reportlab.graphics.charts.piecharts import Pie
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    st.warning("PDF功能需要安装reportlab: pip install reportlab")

# 图表生成
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')  # 使用非交互式后端
import seaborn as sns
from matplotlib.backends.backend_pdf import PdfPages

# 设置中文字体
plt.rcParams['font.sans-serif'] = ['SimHei', 'Arial Unicode MS', 'DejaVu Sans']
plt.rcParams['axes.unicode_minus'] = False


class ExportFormat:
    """导出格式枚举"""
    CSV = "csv"
    EXCEL = "xlsx"
    PDF = "pdf"
    JSON = "json"
    HTML = "html"


class ExportUtils:
    """导出工具类"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
    
    def create_export_interface(self, df: pd.DataFrame, filename_prefix: str = "export") -> None:
        """创建导出界面"""
        if df.empty:
            st.warning("没有数据可供导出")
            return
        
        st.subheader("📥 数据导出")
        
        # 导出选项
        export_col1, export_col2, export_col3 = st.columns(3)
        
        with export_col1:
            export_format = st.selectbox(
                "导出格式",
                options=["Excel (推荐)", "CSV", "PDF报告", "JSON", "HTML"],
                help="选择导出的文件格式"
            )
        
        with export_col2:
            include_charts = st.checkbox(
                "包含图表",
                value=True,
                help="在导出文件中包含数据可视化图表"
            )
        
        with export_col3:
            include_summary = st.checkbox(
                "包含统计摘要",
                value=True,
                help="在导出文件中包含数据统计摘要"
            )
        
        # 高级选项
        with st.expander("🔧 高级导出选项", expanded=False):
            self._render_advanced_export_options()
        
        # 导出按钮
        if st.button("📥 开始导出", type="primary", use_container_width=True):
            self._handle_export_request(
                df=df,
                export_format=export_format,
                filename_prefix=filename_prefix,
                include_charts=include_charts,
                include_summary=include_summary
            )
    
    def _render_advanced_export_options(self):
        """渲染高级导出选项"""
        col1, col2 = st.columns(2)
        
        with col1:
            st.session_state.export_max_rows = st.number_input(
                "最大导出行数",
                min_value=100,
                max_value=100000,
                value=st.session_state.get('export_max_rows', 10000),
                step=1000,
                help="限制导出的最大行数"
            )
            
            st.session_state.export_precision = st.number_input(
                "数值精度",
                min_value=2,
                max_value=10,
                value=st.session_state.get('export_precision', 4),
                help="数值类型的小数位数"
            )
        
        with col2:
            st.session_state.export_date_format = st.selectbox(
                "日期格式",
                options=["%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"],
                index=0,
                help="日期时间的显示格式"
            )
            
            st.session_state.export_encoding = st.selectbox(
                "字符编码",
                options=["utf-8", "gbk", "ascii"],
                index=0,
                help="文件的字符编码格式"
            )
    
    def _handle_export_request(
        self,
        df: pd.DataFrame,
        export_format: str,
        filename_prefix: str,
        include_charts: bool,
        include_summary: bool
    ):
        """处理导出请求"""
        try:
            # 限制导出行数
            max_rows = st.session_state.get('export_max_rows', 10000)
            if len(df) > max_rows:
                df_export = df.head(max_rows)
                st.warning(f"数据量过大，只导出前 {max_rows} 行")
            else:
                df_export = df.copy()
            
            # 格式化数据
            df_export = self._format_export_data(df_export)
            
            # 生成文件名
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{filename_prefix}_{timestamp}"
            
            # 根据格式执行导出
            if export_format == "Excel (推荐)":
                self._export_to_excel(df_export, filename, include_charts, include_summary)
            elif export_format == "CSV":
                self._export_to_csv(df_export, filename)
            elif export_format == "PDF报告":
                if PDF_AVAILABLE:
                    self._export_to_pdf(df_export, filename, include_charts, include_summary)
                else:
                    st.error("PDF功能不可用，请安装reportlab库")
            elif export_format == "JSON":
                self._export_to_json(df_export, filename)
            elif export_format == "HTML":
                self._export_to_html(df_export, filename, include_charts, include_summary)
        
        except Exception as e:
            self.logger.error(f"导出失败: {e}", exc_info=True)
            st.error(f"导出失败: {str(e)}")
    
    def _format_export_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """格式化导出数据"""
        df_formatted = df.copy()
        
        # 数值精度处理
        precision = st.session_state.get('export_precision', 4)
        numeric_columns = df_formatted.select_dtypes(include=[np.number]).columns
        
        for col in numeric_columns:
            if col in ['profit_margin', 'risk_score', 'confidence_score']:
                df_formatted[col] = df_formatted[col].round(precision)
        
        # 日期格式处理
        date_format = st.session_state.get('export_date_format', '%Y-%m-%d %H:%M:%S')
        date_columns = df_formatted.select_dtypes(include=['datetime64']).columns
        
        for col in date_columns:
            df_formatted[col] = df_formatted[col].dt.strftime(date_format)
        
        return df_formatted
    
    def _export_to_excel(
        self,
        df: pd.DataFrame,
        filename: str,
        include_charts: bool,
        include_summary: bool
    ):
        """导出到Excel格式"""
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # 主数据工作表
            df.to_excel(writer, sheet_name='数据', index=False)
            
            # 获取工作簿和工作表
            workbook = writer.book
            worksheet = writer.sheets['数据']
            
            # 设置表格样式
            self._style_excel_worksheet(worksheet, df)
            
            # 添加统计摘要工作表
            if include_summary:
                summary_df = self._generate_summary_statistics(df)
                summary_df.to_excel(writer, sheet_name='统计摘要', index=True)
                self._style_excel_worksheet(writer.sheets['统计摘要'], summary_df)
            
            # 添加图表工作表
            if include_charts and self._has_numeric_data(df):
                self._add_excel_charts(workbook, df)
        
        output.seek(0)
        
        st.download_button(
            label=f"📥 下载 {filename}.xlsx",
            data=output.getvalue(),
            file_name=f"{filename}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        st.success("Excel文件生成成功！")
    
    def _style_excel_worksheet(self, worksheet, df: pd.DataFrame):
        """设置Excel工作表样式"""
        # 标题行样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 边框样式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 应用标题样式
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 数据行样式
        for row in range(2, len(df) + 2):
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = thin_border
                
                # 数值列右对齐
                if col in df.select_dtypes(include=[np.number]).columns:
                    cell.alignment = Alignment(horizontal="right")
        
        # 自动调整列宽
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _add_excel_charts(self, workbook: Workbook, df: pd.DataFrame):
        """添加Excel图表"""
        # 创建图表工作表
        chart_sheet = workbook.create_sheet("图表")
        
        numeric_columns = df.select_dtypes(include=[np.number]).columns.tolist()
        
        if len(numeric_columns) >= 2:
            # 散点图
            chart = ScatterChart()
            chart.title = "数据散点图"
            chart.style = 13
            chart.x_axis.title = numeric_columns[0]
            chart.y_axis.title = numeric_columns[1]
            
            # 添加数据
            data = Reference(workbook['数据'], min_col=2, min_row=1, max_row=min(100, len(df) + 1), max_col=3)
            chart.add_data(data, titles_from_data=True)
            
            chart_sheet.add_chart(chart, "A1")
        
        # 如果有分类列，添加饼图
        if 'strategy_type' in df.columns:
            strategy_counts = df['strategy_type'].value_counts()
            
            # 创建数据表
            start_row = 15
            chart_sheet.cell(row=start_row, column=1, value="策略类型")
            chart_sheet.cell(row=start_row, column=2, value="数量")
            
            for i, (strategy, count) in enumerate(strategy_counts.items(), 1):
                chart_sheet.cell(row=start_row + i, column=1, value=strategy)
                chart_sheet.cell(row=start_row + i, column=2, value=count)
            
            # 创建饼图
            pie_chart = PieChart()
            pie_chart.title = "策略类型分布"
            
            data = Reference(chart_sheet, min_col=2, min_row=start_row + 1, max_row=start_row + len(strategy_counts))
            categories = Reference(chart_sheet, min_col=1, min_row=start_row + 1, max_row=start_row + len(strategy_counts))
            
            pie_chart.add_data(data)
            pie_chart.set_categories(categories)
            
            chart_sheet.add_chart(pie_chart, "H1")
    
    def _export_to_csv(self, df: pd.DataFrame, filename: str):
        """导出到CSV格式"""
        encoding = st.session_state.get('export_encoding', 'utf-8')
        
        csv_data = df.to_csv(index=False, encoding=encoding)
        
        st.download_button(
            label=f"📥 下载 {filename}.csv",
            data=csv_data,
            file_name=f"{filename}.csv",
            mime="text/csv"
        )
        
        st.success("CSV文件生成成功！")
    
    def _export_to_pdf(
        self,
        df: pd.DataFrame,
        filename: str,
        include_charts: bool,
        include_summary: bool
    ):
        """导出到PDF报告"""
        if not PDF_AVAILABLE:
            st.error("PDF功能不可用")
            return
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        styles = getSampleStyleSheet()
        
        # 自定义样式
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # 居中
        )
        
        # 标题
        title = Paragraph("套利机会分析报告", title_style)
        story.append(title)
        story.append(Spacer(1, 12))
        
        # 生成时间
        timestamp = datetime.now().strftime("%Y年%m月%d日 %H:%M:%S")
        story.append(Paragraph(f"生成时间: {timestamp}", styles['Normal']))
        story.append(Spacer(1, 12))
        
        # 统计摘要
        if include_summary:
            story.append(Paragraph("数据摘要", styles['Heading2']))
            
            # 基本统计
            summary_data = [
                ["项目", "数值"],
                ["总记录数", f"{len(df):,}"],
                ["数据列数", f"{len(df.columns)}"],
            ]
            
            if 'profit_margin' in df.columns:
                summary_data.extend([
                    ["平均利润率", f"{df['profit_margin'].mean():.4f}"],
                    ["最大利润率", f"{df['profit_margin'].max():.4f}"],
                ])
            
            if 'risk_score' in df.columns:
                summary_data.extend([
                    ["平均风险评分", f"{df['risk_score'].mean():.4f}"],
                    ["最低风险评分", f"{df['risk_score'].min():.4f}"],
                ])
            
            summary_table = Table(summary_data)
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(summary_table)
            story.append(PageBreak())
        
        # 数据表格
        story.append(Paragraph("详细数据", styles['Heading2']))
        
        # 限制显示的行数和列数（PDF空间有限）
        max_rows = 50
        max_cols = 8
        
        display_df = df.head(max_rows).iloc[:, :max_cols]
        
        # 准备表格数据
        table_data = [list(display_df.columns)]
        
        for _, row in display_df.iterrows():
            formatted_row = []
            for val in row:
                if pd.isna(val):
                    formatted_row.append("N/A")
                elif isinstance(val, float):
                    formatted_row.append(f"{val:.4f}")
                else:
                    formatted_row.append(str(val)[:20])  # 限制字符长度
            table_data.append(formatted_row)
        
        # 创建表格
        data_table = Table(table_data)
        data_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(data_table)
        
        # 图表
        if include_charts and self._has_numeric_data(df):
            story.append(PageBreak())
            story.append(Paragraph("数据可视化", styles['Heading2']))
            
            # 生成图表并添加到PDF
            chart_images = self._generate_charts_for_pdf(df)
            for chart_image in chart_images:
                story.append(chart_image)
                story.append(Spacer(1, 12))
        
        # 生成PDF
        doc.build(story)
        buffer.seek(0)
        
        st.download_button(
            label=f"📥 下载 {filename}.pdf",
            data=buffer.getvalue(),
            file_name=f"{filename}.pdf",
            mime="application/pdf"
        )
        
        st.success("PDF报告生成成功！")
    
    def _generate_charts_for_pdf(self, df: pd.DataFrame) -> List:
        """为PDF生成图表"""
        charts = []
        
        # 设置图表样式
        plt.style.use('default')
        
        numeric_columns = df.select_dtypes(include=[np.number]).columns.tolist()
        
        if len(numeric_columns) >= 2:
            # 散点图
            fig, ax = plt.subplots(figsize=(8, 6))
            ax.scatter(df[numeric_columns[0]], df[numeric_columns[1]], alpha=0.6)
            ax.set_xlabel(numeric_columns[0])
            ax.set_ylabel(numeric_columns[1])
            ax.set_title(f'{numeric_columns[0]} vs {numeric_columns[1]}')
            ax.grid(True, alpha=0.3)
            
            # 保存到临时文件
            img_buffer = io.BytesIO()
            plt.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight')
            img_buffer.seek(0)
            
            # 创建ReportLab图像对象
            img = RLImage(img_buffer, width=400, height=300)
            charts.append(img)
            
            plt.close()
        
        # 如果有分类数据，生成条形图
        if 'strategy_type' in df.columns:
            strategy_counts = df['strategy_type'].value_counts()
            
            fig, ax = plt.subplots(figsize=(8, 6))
            strategy_counts.plot(kind='bar', ax=ax)
            ax.set_title('策略类型分布')
            ax.set_xlabel('策略类型')
            ax.set_ylabel('数量')
            plt.xticks(rotation=45)
            
            img_buffer = io.BytesIO()
            plt.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight')
            img_buffer.seek(0)
            
            img = RLImage(img_buffer, width=400, height=300)
            charts.append(img)
            
            plt.close()
        
        return charts
    
    def _export_to_json(self, df: pd.DataFrame, filename: str):
        """导出到JSON格式"""
        # 处理特殊数据类型
        df_json = df.copy()
        
        # 转换日期时间
        for col in df_json.select_dtypes(include=['datetime64']).columns:
            df_json[col] = df_json[col].dt.strftime('%Y-%m-%d %H:%M:%S')
        
        # 处理NaN值
        df_json = df_json.fillna('')
        
        # 转换为JSON
        json_data = df_json.to_json(orient='records', indent=2, ensure_ascii=False)
        
        st.download_button(
            label=f"📥 下载 {filename}.json",
            data=json_data,
            file_name=f"{filename}.json",
            mime="application/json"
        )
        
        st.success("JSON文件生成成功！")
    
    def _export_to_html(
        self,
        df: pd.DataFrame,
        filename: str,
        include_charts: bool,
        include_summary: bool
    ):
        """导出到HTML格式"""
        html_content = f"""
        <!DOCTYPE html>
        <html lang="zh-CN">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>套利机会分析报告</title>
            <style>
                body {{
                    font-family: 'Arial', sans-serif;
                    margin: 40px;
                    background-color: #f5f5f5;
                }}
                .container {{
                    max-width: 1200px;
                    margin: 0 auto;
                    background-color: white;
                    padding: 30px;
                    border-radius: 10px;
                    box-shadow: 0 0 10px rgba(0,0,0,0.1);
                }}
                h1 {{
                    color: #2c3e50;
                    text-align: center;
                    margin-bottom: 30px;
                }}
                h2 {{
                    color: #34495e;
                    border-bottom: 2px solid #3498db;
                    padding-bottom: 10px;
                }}
                table {{
                    width: 100%;
                    border-collapse: collapse;
                    margin: 20px 0;
                }}
                th, td {{
                    border: 1px solid #ddd;
                    padding: 12px;
                    text-align: left;
                }}
                th {{
                    background-color: #3498db;
                    color: white;
                    font-weight: bold;
                }}
                tr:nth-child(even) {{
                    background-color: #f2f2f2;
                }}
                .summary-box {{
                    background-color: #ecf0f1;
                    padding: 20px;
                    border-radius: 5px;
                    margin: 20px 0;
                }}
                .timestamp {{
                    text-align: center;
                    color: #7f8c8d;
                    font-style: italic;
                    margin-bottom: 30px;
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <h1>套利机会分析报告</h1>
                <p class="timestamp">生成时间: {datetime.now().strftime('%Y年%m月%d日 %H:%M:%S')}</p>
        """
        
        # 添加统计摘要
        if include_summary:
            summary_stats = self._generate_summary_statistics(df)
            html_content += """
                <h2>数据摘要</h2>
                <div class="summary-box">
            """
            html_content += summary_stats.to_html(classes="summary-table")
            html_content += "</div>"
        
        # 添加主数据表
        html_content += "<h2>详细数据</h2>"
        
        # 限制显示行数
        display_df = df.head(1000)  # HTML可以显示更多行
        html_content += display_df.to_html(classes="data-table", table_id="main-table")
        
        html_content += """
            </div>
        </body>
        </html>
        """
        
        st.download_button(
            label=f"📥 下载 {filename}.html",
            data=html_content,
            file_name=f"{filename}.html",
            mime="text/html"
        )
        
        st.success("HTML文件生成成功！")
    
    def _generate_summary_statistics(self, df: pd.DataFrame) -> pd.DataFrame:
        """生成统计摘要"""
        summary_data = {}
        
        # 基本信息
        summary_data['数据概览'] = {
            '总行数': len(df),
            '总列数': len(df.columns),
            '数值列数': len(df.select_dtypes(include=[np.number]).columns),
            '文本列数': len(df.select_dtypes(include=['object']).columns),
            '日期列数': len(df.select_dtypes(include=['datetime64']).columns)
        }
        
        # 数值列统计
        numeric_cols = df.select_dtypes(include=[np.number]).columns
        if len(numeric_cols) > 0:
            numeric_summary = {}
            for col in numeric_cols:
                numeric_summary[f'{col}_均值'] = df[col].mean()
                numeric_summary[f'{col}_中位数'] = df[col].median()
                numeric_summary[f'{col}_标准差'] = df[col].std()
                numeric_summary[f'{col}_最小值'] = df[col].min()
                numeric_summary[f'{col}_最大值'] = df[col].max()
            
            summary_data['数值统计'] = numeric_summary
        
        # 分类列统计
        if 'strategy_type' in df.columns:
            strategy_summary = {}
            strategy_counts = df['strategy_type'].value_counts()
            for strategy, count in strategy_counts.items():
                strategy_summary[f'{strategy}'] = count
            
            summary_data['策略分布'] = strategy_summary
        
        # 转换为DataFrame
        summary_list = []
        for category, stats in summary_data.items():
            for key, value in stats.items():
                summary_list.append({
                    '类别': category,
                    '指标': key,
                    '数值': f"{value:.4f}" if isinstance(value, float) else value
                })
        
        return pd.DataFrame(summary_list)
    
    def _has_numeric_data(self, df: pd.DataFrame) -> bool:
        """检查是否有数值数据可以生成图表"""
        numeric_cols = df.select_dtypes(include=[np.number]).columns
        return len(numeric_cols) >= 1
    
    def create_batch_export_interface(self, datasets: Dict[str, pd.DataFrame]) -> None:
        """创建批量导出界面"""
        if not datasets:
            st.warning("没有数据集可供导出")
            return
        
        st.subheader("📦 批量导出")
        
        # 数据集选择
        selected_datasets = st.multiselect(
            "选择要导出的数据集",
            options=list(datasets.keys()),
            default=list(datasets.keys()),
            help="选择要包含在导出中的数据集"
        )
        
        if not selected_datasets:
            st.warning("请至少选择一个数据集")
            return
        
        # 导出选项
        batch_col1, batch_col2 = st.columns(2)
        
        with batch_col1:
            batch_format = st.selectbox(
                "批量导出格式",
                options=["Excel工作簿", "ZIP压缩包"],
                help="Excel工作簿: 多个工作表; ZIP: 多个独立文件"
            )
        
        with batch_col2:
            include_index = st.checkbox(
                "生成索引文件",
                value=True,
                help="生成包含所有数据集信息的索引文件"
            )
        
        # 批量导出按钮
        if st.button("📦 开始批量导出", type="primary", use_container_width=True):
            if batch_format == "Excel工作簿":
                self._export_batch_excel(datasets, selected_datasets, include_index)
            else:
                self._export_batch_zip(datasets, selected_datasets, include_index)
    
    def _export_batch_excel(
        self,
        datasets: Dict[str, pd.DataFrame],
        selected_datasets: List[str],
        include_index: bool
    ):
        """批量导出到Excel工作簿"""
        output = io.BytesIO()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # 导出每个数据集
            for dataset_name in selected_datasets:
                df = datasets[dataset_name]
                sheet_name = dataset_name[:31]  # Excel工作表名称限制
                
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # 应用样式
                worksheet = writer.sheets[sheet_name]
                self._style_excel_worksheet(worksheet, df)
            
            # 生成索引文件
            if include_index:
                index_data = []
                for dataset_name in selected_datasets:
                    df = datasets[dataset_name]
                    index_data.append({
                        '数据集名称': dataset_name,
                        '行数': len(df),
                        '列数': len(df.columns),
                        '工作表名称': dataset_name[:31],
                        '最后更新': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    })
                
                index_df = pd.DataFrame(index_data)
                index_df.to_excel(writer, sheet_name='索引', index=False)
                self._style_excel_worksheet(writer.sheets['索引'], index_df)
        
        output.seek(0)
        
        st.download_button(
            label=f"📥 下载批量导出_{timestamp}.xlsx",
            data=output.getvalue(),
            file_name=f"批量导出_{timestamp}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        st.success("批量Excel导出完成！")
    
    def _export_batch_zip(
        self,
        datasets: Dict[str, pd.DataFrame],
        selected_datasets: List[str],
        include_index: bool
    ):
        """批量导出到ZIP文件"""
        zip_buffer = io.BytesIO()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # 导出每个数据集
            for dataset_name in selected_datasets:
                df = datasets[dataset_name]
                
                # CSV格式
                csv_buffer = io.StringIO()
                df.to_csv(csv_buffer, index=False, encoding='utf-8')
                zip_file.writestr(f"{dataset_name}.csv", csv_buffer.getvalue())
                
                # JSON格式
                json_data = df.to_json(orient='records', indent=2, ensure_ascii=False)
                zip_file.writestr(f"{dataset_name}.json", json_data)
            
            # 生成索引文件
            if include_index:
                index_data = {
                    'export_info': {
                        'export_time': datetime.now().isoformat(),
                        'total_datasets': len(selected_datasets),
                        'formats': ['CSV', 'JSON']
                    },
                    'datasets': []
                }
                
                for dataset_name in selected_datasets:
                    df = datasets[dataset_name]
                    index_data['datasets'].append({
                        'name': dataset_name,
                        'rows': len(df),
                        'columns': len(df.columns),
                        'files': [f"{dataset_name}.csv", f"{dataset_name}.json"]
                    })
                
                import json
                zip_file.writestr('index.json', json.dumps(index_data, indent=2, ensure_ascii=False))
        
        zip_buffer.seek(0)
        
        st.download_button(
            label=f"📥 下载批量导出_{timestamp}.zip",
            data=zip_buffer.getvalue(),
            file_name=f"批量导出_{timestamp}.zip",
            mime="application/zip"
        )
        
        st.success("批量ZIP导出完成！")
    
    def create_template_export(self) -> None:
        """创建导入模板导出"""
        st.subheader("📄 导入模板")
        
        st.write("下载标准的数据导入模板，用于批量导入套利机会数据")
        
        # 创建模板数据
        template_data = {
            'id': ['OPP_001', 'OPP_002', 'OPP_003'],
            'strategy_type': ['covered_call', 'protective_put', 'iron_condor'],
            'profit_margin': [0.025, 0.018, 0.032],
            'expected_profit': [250.0, 180.0, 320.0],
            'risk_score': [0.15, 0.22, 0.18],
            'confidence_score': [0.85, 0.78, 0.90],
            'instruments': ['AAPL_CALL_150', 'GOOGL_PUT_2500', 'TSLA_IRON_800'],
            'timestamp': [
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            ]
        }
        
        template_df = pd.DataFrame(template_data)
        
        # 显示模板预览
        st.write("**模板预览:**")
        st.dataframe(template_df, use_container_width=True)
        
        # 模板说明
        with st.expander("📖 字段说明", expanded=False):
            field_descriptions = {
                'id': '机会唯一标识符',
                'strategy_type': '策略类型 (covered_call, protective_put, iron_condor等)',
                'profit_margin': '利润率 (小数形式, 如0.025表示2.5%)',
                'expected_profit': '预期利润金额',
                'risk_score': '风险评分 (0-1之间)',
                'confidence_score': '置信度 (0-1之间)',
                'instruments': '相关金融工具',
                'timestamp': '时间戳 (YYYY-MM-DD HH:MM:SS格式)'
            }
            
            for field, description in field_descriptions.items():
                st.write(f"**{field}**: {description}")
        
        # 导出模板
        template_formats = st.multiselect(
            "选择模板格式",
            options=["Excel", "CSV"],
            default=["Excel", "CSV"]
        )
        
        if st.button("📥 下载导入模板", type="primary"):
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            if "Excel" in template_formats:
                excel_buffer = io.BytesIO()
                with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                    template_df.to_excel(writer, sheet_name='数据模板', index=False)
                    
                    # 添加说明工作表
                    desc_data = pd.DataFrame(list(field_descriptions.items()), 
                                           columns=['字段名', '说明'])
                    desc_data.to_excel(writer, sheet_name='字段说明', index=False)
                
                excel_buffer.seek(0)
                
                st.download_button(
                    label="📥 下载Excel模板",
                    data=excel_buffer.getvalue(),
                    file_name=f"导入模板_{timestamp}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            
            if "CSV" in template_formats:
                csv_data = template_df.to_csv(index=False, encoding='utf-8')
                
                st.download_button(
                    label="📥 下载CSV模板",
                    data=csv_data,
                    file_name=f"导入模板_{timestamp}.csv",
                    mime="text/csv"
                )
            
            st.success("模板下载完成！")